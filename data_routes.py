import csv
import io
import sqlite3
from datetime import datetime
from flask import Blueprint, request, session, Response, jsonify
from openpyxl import load_workbook, Workbook

data_bp = Blueprint('data_bp', __name__)

def get_db_connection():
    conn = sqlite3.connect('library_v3.db')
    conn.row_factory = sqlite3.Row
    return conn

# ----------------- TEMPLATES -----------------
@data_bp.route('/template/<module>')
def download_template(module):
    if not session.get('role') in ['admin', 'super_admin']: return jsonify({'error': 'Unauthorized'}), 403
    
    templates = {
        'books': ['title', 'author', 'isbn', 'category', 'quantity', 'publisher', 'description', 'shelf_location', 'school_code'],
        'students': ['studentId', 'name', 'phone', 'class', 'school_code', 'password'],
        'librarians': ['name', 'phone', 'school_code', 'password'],
        'schools': ['name', 'school_code', 'librarian_name']
    }
    
    if module not in templates: return "Invalid module", 400
    
    def generate():
        yield ','.join(templates[module]) + '\n'
        
    return Response(generate(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={module}_template.csv'})

# ----------------- EXPORTS -----------------
@data_bp.route('/export/<module>')
def export_data(module):
    if not session.get('role') in ['admin', 'super_admin']: return jsonify({'error': 'Unauthorized'}), 403
    
    fmt = request.args.get('format', 'csv')
    s_code = request.args.get('school', session.get('school_code'))
    
    # Security check: Admins can only export their own school
    if session.get('role') == 'admin':
        s_code = session.get('school_code')
        
    conn = get_db_connection()
    data = []
    headers = []
    
    if module == 'books':
        headers = ['Title', 'Author', 'ISBN', 'Category', 'Total Copies', 'Available', 'School Code', 'School Name', 'Librarian Name']
        query = '''SELECT b.title, b.author, b.barcode_id, b.genre, b.total_copies, b.available_copies, 
                          b.school_code, s.name as school_name, s.librarian_name 
                   FROM books b LEFT JOIN schools s ON b.school_code = s.school_code'''
        if s_code and s_code != 'ALL':
            data = conn.execute(query + ' WHERE b.school_code = ?', (s_code,)).fetchall()
        else:
            data = conn.execute(query).fetchall()
            
    elif module == 'students':
        headers = ['Student ID', 'Name', 'Phone (Login ID)', 'Class', 'School Code', 'Password', 'School Name', 'Librarian Name', 'Status']
        query = '''SELECT u.admission_no, u.name, u.phone, u.class, u.school_code, u.password,
                          s.name as school_name, s.librarian_name, u.status 
                   FROM users u LEFT JOIN schools s ON u.school_code = s.school_code 
                   WHERE u.role="student"'''
        if s_code and s_code != 'ALL':
            data = conn.execute(query + ' AND u.school_code = ?', (s_code,)).fetchall()
        else:
            data = conn.execute(query).fetchall()
            
    elif module == 'transactions':
        headers = ['Student Name', 'Book Title', 'School Code', 'Issue Date', 'Due Date', 'Return Date', 'Fine']
        query = '''SELECT u.name as student, b.title as book, t.school_code, t.issue_date, t.due_date, t.return_date, t.fine 
                   FROM transactions t 
                   JOIN users u ON t.user_id = u.id 
                   JOIN books b ON t.book_id = b.id'''
        if s_code and s_code != 'ALL':
            data = conn.execute(query + ' WHERE t.school_code = ?', (s_code,)).fetchall()
        else:
            data = conn.execute(query).fetchall()

    elif module == 'librarians':
        headers = ['Name', 'Phone', 'School Code', 'Password', 'Status']
        query = '''SELECT name, phone, school_code, password, status FROM users WHERE role="admin"'''
        if s_code and s_code != 'ALL':
            data = conn.execute(query + ' AND school_code = ?', (s_code,)).fetchall()
        else:
            data = conn.execute(query).fetchall()

    elif module == 'schools':
        headers = ['ID', 'Name', 'School Code', 'Librarian Name', 'Status', 'Created At']
        query = '''SELECT id, name, school_code, librarian_name, status, created_at FROM schools'''
        if s_code and s_code != 'ALL':
            data = conn.execute(query + ' WHERE school_code = ?', (s_code,)).fetchall()
        else:
            data = conn.execute(query).fetchall()
    
    conn.execute('INSERT INTO logs (user_id, action, module, created_at, school_code) VALUES (?, ?, ?, ?, ?)',
                 (session.get('user_id'), f"Exported {module} ({fmt})", "Export", datetime.now().strftime('%Y-%m-%d %H:%M'), s_code))
    conn.commit()
    conn.close()
    
    if fmt == 'csv':
        def generate():
            yield ','.join(headers) + '\n'
            for row in data:
                yield ','.join([str(val) if val is not None else '' for val in row]) + '\n'
        return Response(generate(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={module}_export.csv'})
        
    elif fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data:
            ws.append([val for val in row])
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return Response(out.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={module}_export.xlsx'})

# ----------------- IMPORTS -----------------
@data_bp.route('/import/<module>', methods=['POST'])
def import_data(module):
    if not session.get('role') in ['admin', 'super_admin']: return jsonify({'error': 'Unauthorized'}), 403
    
    if 'file' not in request.files: return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename: return jsonify({'error': 'Empty filename'}), 400
    
    s_code = request.form.get('school', session.get('school_code'))
    if session.get('role') == 'admin':
        s_code = session.get('school_code')
        
    if not s_code or s_code == 'ALL':
        return jsonify({'error': 'Must select a specific target school.'}), 400

    filename = file.filename.lower()
    records = []
    
    # Parse File
    if filename.endswith('.csv'):
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_input = csv.DictReader(stream)
        for row in csv_input:
            # Clean keys (lowercase, strip whitespace)
            clean_row = {k.strip().lower(): v.strip() for k, v in row.items() if k and v}
            if clean_row: records.append(clean_row)
            
    elif filename.endswith('.xlsx'):
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h).strip().lower() for h in rows[0] if h]
            for row in rows[1:]:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and val is not None:
                        row_dict[headers[i]] = str(val).strip()
                if row_dict: records.append(row_dict)
    else:
        return jsonify({'error': 'Unsupported file format. Use .csv or .xlsx'}), 400

    if not records:
        return jsonify({'error': 'File is empty or invalid.'}), 400

    conn = get_db_connection()
    success = 0
    failed = 0
    duplicates = 0
    errors = []

    # Batch Process
    for index, row in enumerate(records):
        try:
            if module == 'books':
                title = row.get('title')
                author = row.get('author')
                isbn = row.get('isbn')
                category = row.get('category', 'General')
                qty = int(row.get('quantity', 0))
                
                if not title or not author or not isbn or qty < 1:
                    failed += 1
                    errors.append(f"Row {index+1}: Missing required book fields")
                    continue
                    
                target_school = row.get('school_code') or s_code
                if not target_school or target_school == 'ALL':
                    failed += 1
                    errors.append(f"Row {index+1}: Missing target school")
                    continue

                # Check duplicate ISBN
                exists = conn.execute('SELECT id FROM books WHERE barcode_id = ? AND school_code = ?', (isbn, target_school)).fetchone()
                if exists:
                    duplicates += 1
                    continue
                    
                conn.execute('''INSERT INTO books (title, author, barcode_id, genre, total_copies, available_copies, school_code, description, shelf_location) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                             (title, author, isbn, category, qty, qty, target_school, row.get('description', ''), row.get('shelf_location', '')))
                success += 1

            elif module == 'students':
                sid = row.get('studentid')
                name = row.get('name')
                phone = row.get('phone')
                cls = row.get('class')
                password = row.get('password', 'studentpass')
                
                if not sid or not name or not phone or not cls:
                    failed += 1
                    errors.append(f"Row {index+1}: Missing required student fields")
                    continue
                    
                target_school = row.get('school_code') or s_code
                if not target_school or target_school == 'ALL':
                    failed += 1
                    errors.append(f"Row {index+1}: Missing target school")
                    continue

                exists = conn.execute('SELECT id FROM users WHERE phone = ?', (phone,)).fetchone()
                if exists:
                    duplicates += 1
                    continue
                    
                conn.execute('''INSERT INTO users (name, admission_no, phone, class, role, password, school_code, status) 
                                VALUES (?, ?, ?, ?, 'student', ?, ?, 'active')''',
                             (name, sid, phone, cls, password, target_school))
                success += 1

            elif module == 'librarians':
                name = row.get('name')
                phone = row.get('phone')
                password = row.get('password', 'adminpass')
                
                if not name or not phone:
                    failed += 1
                    errors.append(f"Row {index+1}: Missing required librarian fields")
                    continue
                    
                target_school = row.get('school_code') or s_code
                if not target_school or target_school == 'ALL':
                    failed += 1
                    errors.append(f"Row {index+1}: Missing target school")
                    continue

                exists = conn.execute('SELECT id FROM users WHERE phone = ?', (phone,)).fetchone()
                if exists:
                    duplicates += 1
                    continue
                    
                conn.execute('''INSERT INTO users (name, phone, role, password, school_code, status) 
                                VALUES (?, ?, 'admin', ?, ?, 'active')''',
                             (name, phone, password, target_school))
                success += 1

            elif module == 'schools':
                name = row.get('name')
                school_code = row.get('school_code')
                librarian_name = row.get('librarian_name', '')
                
                if not name or not school_code:
                    failed += 1
                    errors.append(f"Row {index+1}: Missing required school fields")
                    continue
                
                exists = conn.execute('SELECT id FROM schools WHERE school_code = ?', (school_code,)).fetchone()
                if exists:
                    duplicates += 1
                    continue
                    
                conn.execute('''INSERT INTO schools (name, school_code, librarian_name, created_at, status) 
                                VALUES (?, ?, ?, ?, 'Active')''',
                             (name, school_code, librarian_name, datetime.now().strftime('%Y-%m-%d %H:%M')))
                success += 1
                
        except Exception as e:
            failed += 1
            errors.append(f"Row {index+1}: {str(e)}")

    # Log action
    conn.execute('INSERT INTO logs (user_id, action, module, created_at, school_code) VALUES (?, ?, ?, ?, ?)',
                 (session.get('user_id'), f"Imported {success} {module}", "Import", datetime.now().strftime('%Y-%m-%d %H:%M'), s_code))
    conn.commit()
    conn.close()

    return jsonify({
        'total': len(records),
        'success': success,
        'failed': failed,
        'duplicates': duplicates,
        'errors': errors[:10] # Return max 10 errors to avoid huge payloads
    })
